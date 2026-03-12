"""
RoGov-SQL – Faza 1: Curățare date educație
Surse:
  - retea-scolara-2024-2025.xlsx   (header pe rândul 4, index 3)
  - elevi-inmatriculati-2024-2025.xlsx (header pe rândul 1, index 0)
Output:
  - edu_reteaua_scolara.sql  (schema + INSERT-uri)
"""

import openpyxl
from collections import defaultdict

# ──────────────────────────────────────────────
# 0. Dicționar județe (cod → nume complet)
# ──────────────────────────────────────────────
COUNTY_MAP = {
    "AB": "Alba",         "AR": "Arad",          "AG": "Argeș",
    "BC": "Bacău",        "BH": "Bihor",         "BN": "Bistrița-Năsăud",
    "BT": "Botoșani",     "BV": "Brașov",        "BR": "Brăila",
    "B":  "București",    "BZ": "Buzău",         "CS": "Caraș-Severin",
    "CL": "Călărași",     "CJ": "Cluj",          "CT": "Constanța",
    "CV": "Covasna",      "DB": "Dâmbovița",     "DJ": "Dolj",
    "GL": "Galați",       "GR": "Giurgiu",       "GJ": "Gorj",
    "HR": "Harghita",     "HD": "Hunedoara",     "IL": "Ialomița",
    "IS": "Iași",         "IF": "Ilfov",         "MM": "Maramureș",
    "MH": "Mehedinți",    "MS": "Mureș",         "NT": "Neamț",
    "OT": "Olt",          "PH": "Prahova",       "SM": "Satu Mare",
    "SJ": "Sălaj",        "SB": "Sibiu",         "SV": "Suceava",
    "TR": "Teleorman",    "TM": "Timiș",         "TL": "Tulcea",
    "VS": "Vaslui",       "VL": "Vâlcea",        "VN": "Vrancea",
}

def esc(val):
    if val is None:
        return "NULL"
    s = str(val).replace("'", "''")
    return f"'{s}'"

def int_or_null(val):
    try:
        return str(int(val))
    except (TypeError, ValueError):
        return "NULL"

# ──────────────────────────────────────────────
# 1. Citire retea-scolara (header pe rândul 4)
# ──────────────────────────────────────────────
print("Citire retea-scolara-2024-2025.xlsx ...")
wb_r = openpyxl.load_workbook(
    'retea-scolara-2024-2025.xlsx',
    read_only=True, data_only=True
)
ws_r = wb_r.active

retea_rows = []
retea_headers = None
for i, row in enumerate(ws_r.iter_rows(values_only=True)):
    if i == 3:
        retea_headers = [str(c).strip() if c else f"col_{i}" for c, i in zip(row, range(len(row)))]
        continue
    if i < 4:
        continue
    if all(c is None for c in row):
        continue
    retea_rows.append(dict(zip(retea_headers, row)))

print(f"  → {len(retea_rows)} rânduri citite din retea-scolara")

# ──────────────────────────────────────────────
# 2. Citire elevi-inmatriculati (header pe rândul 1)
# ──────────────────────────────────────────────
print("Citire elevi-inmatriculati-2024-2025.xlsx ...")
wb_e = openpyxl.load_workbook(
    'elevi-inmatriculati-2024-2025.xlsx',
    read_only=True, data_only=True
)
ws_e = wb_e.active

elevi_rows = []
elevi_headers = None
for i, row in enumerate(ws_e.iter_rows(values_only=True)):
    if i == 0:
        elevi_headers = [str(c).strip() if c else f"col_{i}" for c, i in zip(row, range(len(row)))]
        continue
    if all(c is None for c in row):
        continue
    elevi_rows.append(dict(zip(elevi_headers, row)))

print(f"  → {len(elevi_rows)} rânduri citite din elevi-inmatriculati")

# ──────────────────────────────────────────────
# 3. Agregate din elevi: per siiir_pj
#    - SUM(Elevi exist anterior-asoc) → number_of_students
#    - niveluri unice sortate         → education_level
# ──────────────────────────────────────────────
print("Agregare date elevi per școală ...")
agg_students = defaultdict(int)
agg_levels   = defaultdict(set)

for row in elevi_rows:
    cod = str(row.get('Cod unitate PJ', '') or '').strip()
    
    if not cod:
        continue

    # Normalizăm codul la 10 caractere pentru a nu pierde zerourile inițiale la mapare
    cod = cod.zfill(10)
    
    elevi = row.get('Elevi exist anterior-asoc')
    try:
        agg_students[cod] += int(elevi)
    except (TypeError, ValueError):
        pass
    nivel = str(row.get('Nivel', '') or '').strip()
    if nivel:
        agg_levels[cod].add(nivel)

# Convertim nivelurile în string sortat
agg_levels_str = {
    cod: ', '.join(sorted(niveluri))
    for cod, niveluri in agg_levels.items()
}

print(f"  → {len(agg_students)} școli cu date de elevi")

# ──────────────────────────────────────────────
# 4. Construire COUNTIES
# ──────────────────────────────────────────────
county_codes_seen = set()
for row in retea_rows:
    cod = str(row.get('Judet PJ', '') or '').strip()
    if cod:
        county_codes_seen.add(cod)

# Sortăm după cod pentru ID consistent
counties = []
for idx, cod in enumerate(sorted(county_codes_seen), start=1):
    counties.append({
        'county_id':   idx,
        'county_code': cod,
        'county_name': COUNTY_MAP.get(cod, cod),
    })

county_code_to_id = {c['county_code']: c['county_id'] for c in counties}
print(f"  → {len(counties)} județe")

# ──────────────────────────────────────────────
# 5. Construire LOCALITIES
#    Cheie de deduplicare: (locality_name_upper, county_id)
# ──────────────────────────────────────────────
locality_key_to_id = {}
localities = []
loc_idx = 1

for row in retea_rows:
    loc_name  = str(row.get('Localitate unitate', '') or '').strip()
    mediu     = str(row.get('Mediu loc. unitate', '') or '').strip()
    county_cod = str(row.get('Judet PJ', '') or '').strip()

    if not loc_name or not county_cod:
        continue

    county_id = county_code_to_id.get(county_cod)
    if county_id is None:
        continue

    key = (loc_name.upper(), county_id)
    if key not in locality_key_to_id:
        locality_key_to_id[key] = loc_idx
        localities.append({
            'locality_id':    loc_idx,
            'locality_name':  loc_name,
            'residency_area': mediu,
            'county_id':      county_id,
        })
        loc_idx += 1

print(f"  → {len(localities)} localități unice")

# ──────────────────────────────────────────────
# 6. Construire SCHOOLS din retea-scolara
#    - Fiecare rând = o unitate (PJ sau AR)
#    - siiir_code = Cod SIIIR unitate
#    - school_name = Denumire PJ (pentru PJ) sau Denumire lunga unitate (pentru AR)
#    - number_of_students / education_level → din agregate (doar PJ au date)
# ──────────────────────────────────────────────
schools = []
seen_siiir = set()
school_idx = 1

for row in retea_rows:
    siiir_unit = str(row.get('Cod SIIIR unitate', '') or '').strip()
    if not siiir_unit:
        continue
        
    siiir_unit = siiir_unit.zfill(10)

    # Evităm încălcarea regulii UNIQUE din baza de date ignorând dublurile
    if siiir_unit in seen_siiir:
        continue
        
    seen_siiir.add(siiir_unit)

    statut = str(row.get('Statut unitate', '') or '').strip()  # PJ sau AR

    # Denumire: pentru PJ folosim Denumire PJ, pentru AR - Denumire lunga unitate
    if statut == 'PJ':
        name = str(row.get('Denumire PJ', '') or '').strip()
    else:
        name = str(row.get('Denumire lunga unitate', '') or '').strip()
        if not name:
            name = str(row.get('Denumire PJ', '') or '').strip()

    short_name    = str(row.get('Denumire scurta unitate', '') or '').strip() or None
    ownership     = str(row.get('Forma proprietate', '') or '').strip() or None
    loc_name      = str(row.get('Localitate unitate', '') or '').strip()
    county_cod    = str(row.get('Judet PJ', '') or '').strip()
    county_id     = county_code_to_id.get(county_cod)

    # Găsim locality_id
    locality_id = None
    if loc_name and county_id:
        locality_id = locality_key_to_id.get((loc_name.upper(), county_id))

    # Legătura dintre cele 2 fișiere xlsx (JOIN logic în memorie)
    # Cheia de legătură: 'Cod SIIIR PJ' (din rețea) == 'Cod unitate PJ' (din agregate elevi)
    siiir_pj = str(row.get('Cod SIIIR PJ', '') or '').strip()
    if siiir_pj:
        siiir_pj = siiir_pj.zfill(10)

    # Extragem agregatele DOAR pentru unitățile cu Personalitate Juridică (PJ).
    # Pentru structurile arondate (AR), raportarea elevilor este centralizată pe PJ-ul de care aparțin, deci lăsăm NULL.
    if statut == 'PJ':
        num_students = agg_students.get(siiir_pj)
        edu_level      = agg_levels_str.get(siiir_pj) or None
    else:
        num_students = None
        edu_level    = None

    schools.append({
        'school_id':        school_idx,
        'school_name':      name,
        'short_name':       short_name,
        'siiir_code':       siiir_unit,
        'locality_id':      locality_id,
        'ownership_type':   ownership,
        'unit_type':        statut,
        'education_level':  edu_level,
        'number_of_students': int(num_students) if num_students is not None else None,
    })
    school_idx += 1

print(f"  → {len(schools)} școli (PJ + AR, deduplicate)")


# ──────────────────────────────────────────────
# 7. Generare SQL
# ──────────────────────────────────────────────
print("Generare edu_reteaua_scolara.sql ...")

lines = []

lines.append("-- ============================================================")
lines.append("-- RoGov-SQL – educatie")
lines.append("-- Generat din: retea-scolara-2024-2025.xlsx & elevi-inmatriculati-2024-2025.xlsx")
lines.append("-- ============================================================\n")

lines.append("PRAGMA foreign_keys = ON;\n")
lines.append("BEGIN TRANSACTION;\n")

# Schema
lines.append("-- ─────────────────── SCHEMA ───────────────────\n")

lines.append("""CREATE TABLE IF NOT EXISTS counties (
    county_id   INT          PRIMARY KEY,
    county_code VARCHAR(10)  NOT NULL UNIQUE,
    county_name VARCHAR(100) NOT NULL
);\n""")

lines.append("""CREATE TABLE IF NOT EXISTS localities (
    locality_id    INT          PRIMARY KEY,
    locality_name  VARCHAR(200) NOT NULL,
    residency_area VARCHAR(50),
    county_id      INT          NOT NULL,
    FOREIGN KEY (county_id) REFERENCES counties(county_id)
);\n""")

lines.append("""CREATE TABLE IF NOT EXISTS schools (
    school_id          INT          PRIMARY KEY,
    school_name        VARCHAR(300) NOT NULL,
    short_name         VARCHAR(100),
    siiir_code         VARCHAR(20)  NOT NULL UNIQUE,
    locality_id        INT,
    ownership_type     VARCHAR(100),
    unit_type          VARCHAR(10),
    education_level    VARCHAR(200),
    number_of_students INT,
    FOREIGN KEY (locality_id) REFERENCES localities(locality_id)
);\n""")

# INSERT counties
lines.append("-- ─────────────────── COUNTIES ─────────────────\n")
lines.append("INSERT INTO counties (county_id, county_code, county_name) VALUES")
county_inserts = []
for c in counties:
    county_inserts.append(
        f"  ({c['county_id']}, {esc(c['county_code'])}, {esc(c['county_name'])})"
    )
lines.append(',\n'.join(county_inserts) + ';\n')

# INSERT localities (în batch-uri de 500)
lines.append("-- ─────────────────── LOCALITIES ────────────────\n")
BATCH = 500
for start in range(0, len(localities), BATCH):
    batch = localities[start:start+BATCH]
    lines.append("INSERT INTO localities (locality_id, locality_name, residency_area, county_id) VALUES")
    loc_inserts = []
    for l in batch:
        loc_inserts.append(
            f"  ({l['locality_id']}, {esc(l['locality_name'])}, {esc(l['residency_area'])}, {l['county_id']})"
        )
    lines.append(',\n'.join(loc_inserts) + ';\n')

# INSERT schools (în batch-uri de 500)
lines.append("-- ─────────────────── SCHOOLS ───────────────────\n")
for start in range(0, len(schools), BATCH):
    batch = schools[start:start+BATCH]
    lines.append("INSERT INTO schools (school_id, school_name, short_name, siiir_code, locality_id, ownership_type, unit_type, education_level, number_of_students) VALUES")
    sch_inserts = []
    for s in batch:
        sch_inserts.append(
            f"  ({s['school_id']}, {esc(s['school_name'])}, {esc(s['short_name'])}, "
            f"{esc(s['siiir_code'])}, {int_or_null(s['locality_id'])}, "
            f"{esc(s['ownership_type'])}, {esc(s['unit_type'])}, "
            f"{esc(s['education_level'])}, {int_or_null(s['number_of_students'])})"
        )
    lines.append(',\n'.join(sch_inserts) + ';\n')

lines.append("COMMIT;\n")

sql_content = '\n'.join(lines)

output_path = 'edu_reteaua_scolara.sql'
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(sql_content)

print(f"\n Fișier generat: {output_path}")
print(f"   Counties  : {len(counties)}")
print(f"   Localities: {len(localities)}")
print(f"   Schools   : {len(schools)}")
print(f"   Dimensiune: {len(sql_content):,} caractere")
